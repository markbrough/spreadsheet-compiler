from flask import abort, Blueprint, render_template, send_from_directory, \
jsonify, request, redirect, url_for, flash, current_app, send_file
from flask_login import login_required, current_user
import sqlalchemy as sa

from spreadsheetcompiler.extensions import login_manager, db
from .. import models
from .users import role_required

import os
import datetime
import openpyxl
from openpyxl import load_workbook


blueprint = Blueprint('compiler', __name__,
    url_prefix='/compiler/', static_folder='../static')


ALLOWED_EXTENSIONS = set(['xlsx'])
def allowed_file(filename):
    return '.' in filename and \
    filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@blueprint.route("/")
@login_required
@role_required(["compiler"])
def compiler():
    if request.args.get('compilation_id'):
        compilation = models.Compilation.query.get(request.args.get('compilation_id'))
        if compilation == None:
            flash("No data found!")
            return redirect(url_for("compiler.compiler"))
    else:
        compilation = active_compilation()
    compilations = models.Compilation.query.order_by(
        models.Compilation.id.desc()).all()
    return render_template("compiler.html",
        compilations=compilations,
        current_compilation = compilation,
        current_user=current_user)


@blueprint.before_app_request
def check_create_compilation():
    active_compilation()


def most_recent_finalised_compilation():
    # Most recent compilation where the final deadline is in the past
    return models.Compilation.query.filter(
        datetime.datetime.utcnow()>models.Compilation.final_deadline
    ).order_by(
        models.Compilation.id.desc()
    ).first()


def active_compilation():
    compilation = models.Compilation.query.filter(
        models.Compilation.created_date<datetime.datetime.utcnow(),
        models.Compilation.final_deadline>datetime.datetime.utcnow()
    ).first()
    if not compilation:
        compilation = models.Compilation()
        templatefile = models.TemplateFile.query.order_by( # Use the highest template file ID
            models.TemplateFile.id.desc()).first()
        compilation.templatefile_id = templatefile.id
        db.session.add(compilation)
        db.session.commit()
    return compilation


def handle_upload(file, set_active_template):
    print("Set active template is {}".format(set_active_template))
    if file.filename == '':
        flash('Please select a file.', "warning")
        return redirect(url_for('compiler.compiler'))
    if file and allowed_file(file.filename):
        template_file = models.TemplateFile(
            user_id=current_user.id
            )
        db.session.add(template_file)
        db.session.commit()
        filename = "{}.xlsx".format(template_file.id)
        path_to_file = os.path.join(current_app.config['UPLOAD_FOLDER'], "templates", filename)
        file.save(path_to_file)
        wb = load_workbook(filename=path_to_file)
        for sheet_number, sheet_name in enumerate(wb.sheetnames):
            template_file.sheets.append(models.TemplateFileSheet(
                name=sheet_name,
                sheet_number=sheet_number))
        db.session.add(template_file)
        db.session.commit()
        _active_compilation = active_compilation()
        if (set_active_template == True) or (_active_compilation.templatefile_id == None):
            _active_compilation.templatefile_id = template_file.id
            db.session.add(_active_compilation)
            db.session.commit()
        flash("Uploaded!", "success")
        return redirect(url_for('compiler.compiler'))

    flash("Please upload an Excel file in the template format.")
    return redirect(url_for('compiler.compiler'))


@blueprint.route("/make/<compilation_id>")
@login_required
@role_required(["compiler", "manager"])
def make(compilation_id=None):
    if not compilation_id:
        compilation = active_compilation()
    else:
        compilation = models.Compilation.query.get(compilation_id)
    sheets = compilation.sheets
    wb = load_workbook(filename=os.path.join(
        current_app.config['UPLOAD_FOLDER'], "templates", "{}.xlsx".format(compilation.templatefile.id)))
    for sheet in sheets:
        _to_ws = wb[sheet.name]
        if not hasattr(sheet, "file"):
            rows = range(3,20)
            cols = range(2, 6)
            flash("WARNING: no data submitted for {}".format(sheet.name))
            for row in rows:
                for col in cols:
                    cell = _to_ws.cell(row=row, column=col)
                    if type(_to_ws[cell.coordinate]) == openpyxl.cell.cell.MergedCell: continue
                    cell.value = "NO DATA SUBMITTED"
            continue
        source = load_workbook(os.path.join(current_app.config['UPLOAD_FOLDER'], "source", sheet.file.filename))
        _from_ws = source[sheet.name]

        for row in _from_ws:
            for cell in row:
                if type(cell) == openpyxl.cell.cell.MergedCell: continue
                if type(_to_ws[cell.coordinate]) == openpyxl.cell.cell.MergedCell: continue
                if hasattr(cell, "value") and hasattr(_to_ws[cell.coordinate], "value"):
                    _to_ws[cell.coordinate].value = cell.value
    wb.save(os.path.join(current_app.config['UPLOAD_FOLDER'], "out.xlsx"))

    attachment_filename = "{} for {}.xlsx".format(
        current_app.config["COMPILED_FILE_NAME"],
        compilation.final_deadline.date().isoformat())
    return send_file(
        os.path.join(current_app.config['UPLOAD_FOLDER'], "out.xlsx"),
        as_attachment=True,
        attachment_filename=attachment_filename,
        cache_timeout=1)


@blueprint.route("/delete_file/<file_id>/")
@login_required
@role_required(["compiler"])
def delete_file(file_id):
    file = models.UploadedFile.query.get(file_id)
    if file and file.user_id == current_user.id:
        os.remove(os.path.join(current_app.config['UPLOAD_FOLDER'], "templates", file.filename))
        db.session.delete(file)
        db.session.commit()
        flash("File deleted.", "success")
        return(redirect(url_for('compiler.compiler')))
    else:
        flash("Could not delete that file.")
        return(redirect(url_for('compiler.compiler')))


@blueprint.route("/template_post/", methods=['GET', 'POST'])
@login_required
@role_required(["compiler"])
def template_post():
    if (request.method == "GET") or ('file' not in request.files):
        flash("You must upload a file!")
        return redirect(url_for('compiler.compiler'))
    file = request.files['file']
    set_active_template = bool(request.form.get('setthisweek') == 'thisweek')
    return handle_upload(file, set_active_template)
