from flask import abort, Blueprint, render_template, send_from_directory, \
jsonify, request, redirect, url_for, flash, abort, current_app, send_file
from flask_login import login_required, current_user

from spreadsheetcompiler.extensions import login_manager, db
from .. import models
import os
import openpyxl
from io import BytesIO
from . import compiler
from .users import role_required

blueprint = Blueprint('data', __name__,
    url_prefix='/data/', static_folder='../static')


@blueprint.route("/download/source/<file_id>/")
@login_required
def download_file(file_id):
    the_file = models.UploadedFile.query.get(file_id)
    if not the_file: return abort(404)
    filename_extension = the_file.filename.rsplit('.', 1)[1].lower()
    attachment_filename = "{} {} - {}.{}".format(
        the_file.user.usergroup.name,
        the_file.created_date.date().isoformat(),
        current_app.config["COMPILED_FILE_NAME"],
        filename_extension)
    return send_file(
        os.path.join(current_app.config['UPLOAD_FOLDER'], "source", the_file.filename),
        as_attachment=True,
        attachment_filename=attachment_filename,
        cache_timeout=1)

@blueprint.route("/download/template/<file_id>/")
@login_required
def download_template(file_id):
    the_file = models.TemplateFile.query.get(file_id)
    if not the_file: return abort(404)
    attachment_filename = "{} Template {}.xlsx".format(
        app.config["COMPILED_FILE_NAME"],
        the_file.created_date.date().isoformat())
    return send_file(
        os.path.join(current_app.config['UPLOAD_FOLDER'], "templates", "{}.xlsx".format(the_file.id)),
        as_attachment=True,
        attachment_filename=attachment_filename,
        cache_timeout=1)

@blueprint.route("/download/uploader_template/")
@login_required
@role_required(["uploader"])
def download_uploader_template():
    compilation = compiler.active_compilation()
    the_file = compilation.templatefile
    if not the_file: return abort(404)
    the_file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], "templates", "{}.xlsx".format(the_file.id))
    attachment_filename = "{} Template {} {}.xlsx".format(
        app.config["COMPILED_FILE_NAME"],
        compilation.final_deadline.date().isoformat(),
        current_user.usergroup.name)
    wb = openpyxl.load_workbook(the_file_path)
    if current_user.usergroup.name not in wb.sheetnames:
        flash("""No sheet found in template for you ({}). Please contact
            the compiler to ensure your sheet exists in the
            template.""".format(current_user.usergroup.name))
        if request.referrer: return redirect(request.referrer)
        return redirect(url_for("users.login"))
    for sheetname in wb.sheetnames:
        if sheetname != current_user.usergroup.name:
            wb.remove(wb.get_sheet_by_name(sheetname))

    the_file = BytesIO()
    wb.save(the_file)
    the_file.seek(0)

    return send_file(
        the_file,
        as_attachment=True,
        attachment_filename=attachment_filename,
        cache_timeout=1)
